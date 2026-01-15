#!/usr/bin/env python3
import openpyxl
import openpyxl.utils as utils
from openpyxl.utils import get_column_letter
import json

# Ouvrir le fichier
wb = openpyxl.load_workbook('test-charges-dual-sections.xlsx')
ws = wb.active

# Lire toutes les lignes
all_rows = []
for row in ws.iter_rows(values_only=True):
    all_rows.append(row)

print(f"📊 Total rows: {len(all_rows)}")
print("\n🔍 All rows content:")
for i, row in enumerate(all_rows):
    print(f"Row {i}: {row}")

# Trouver les sections
print("\n🔎 Looking for section headers...")
avec_tva_idx = -1
sans_tva_idx = -1

for i, row in enumerate(all_rows):
    if row[0]:
        first_cell = str(row[0]).lower()
        print(f"Row {i}, first cell: {first_cell}")
        if "achat avec tva" in first_cell:
            avec_tva_idx = i
            print(f"  ✅ Found 'ACHAT AVEC TVA' at row {i}")
        elif "achat sans tva" in first_cell:
            sans_tva_idx = i
            print(f"  ✅ Found 'ACHAT SANS TVA' at row {i}")

print(f"\n📊 Section indices:")
print(f"  AVEC TVA: row {avec_tva_idx} (data starts at {avec_tva_idx + 2})")
print(f"  SANS TVA: row {sans_tva_idx} (data starts at {sans_tva_idx + 2})")

# Extraire les données avec sheet_to_json
from openpyxl.utils import get_column_letter

# Simuler sheet_to_json behavior
print("\n📊 Simulating sheet_to_json parse:")

# Trouver les rangées d'en-têtes
header_row_avec = avec_tva_idx + 1
header_row_sans = sans_tva_idx + 1

# Extraire les en-têtes AVEC TVA
headers_avec = all_rows[header_row_avec]
print(f"\nHeaders AVEC TVA (row {header_row_avec}): {headers_avec}")

data_avec = []
for i in range(header_row_avec + 1, sans_tva_idx if sans_tva_idx > 0 else len(all_rows)):
    if all_rows[i][0]:  # Si première colonne n'est pas vide
        row_dict = {}
        for j, header in enumerate(headers_avec):
            if header:
                row_dict[header] = all_rows[i][j] if j < len(all_rows[i]) else None
        if any(row_dict.values()):  # Si la ligne n'est pas vide
            data_avec.append(row_dict)

print(f"\nData AVEC TVA ({len(data_avec)} rows):")
for item in data_avec:
    print(f"  {item}")

# Extraire les en-têtes SANS TVA
headers_sans = all_rows[header_row_sans]
print(f"\nHeaders SANS TVA (row {header_row_sans}): {headers_sans}")

data_sans = []
for i in range(header_row_sans + 1, len(all_rows)):
    if all_rows[i][0]:  # Si première colonne n'est pas vide
        row_dict = {}
        for j, header in enumerate(headers_sans):
            if header:
                row_dict[header] = all_rows[i][j] if j < len(all_rows[i]) else None
        if any(row_dict.values()):  # Si la ligne n'est pas vide
            data_sans.append(row_dict)

print(f"\nData SANS TVA ({len(data_sans)} rows):")
for item in data_sans:
    print(f"  {item}")

print("\n✅ Section detection successful!")
